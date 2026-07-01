from fastapi import FastAPI, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, PlainTextResponse
from pydantic import BaseModel
from typing import Optional
import io
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .models import (
    CalculationRequest, CalculationResponse, GridPoint, SourceResult,
    SubstanceResult, TopPoint, TopPointContribution,
    SzzResult, SzzBoundaryPoint,
)
from .meteo_data import CITIES
from .ond86 import compute_grid, compute_szz_boundary, compute_per_substance
from .pdf_export import generate_pdf, _make_transparent_dispersion_map, _make_concentration_plot
from .substances import SUBSTANCES
from .tables import generate_pdv_tables, generate_ovos_tables
from .import_sources import parse_csv, parse_excel, generate_template_csv, generate_template_xlsx


# ---------------------------------------------------------------------------
# Persistence helpers for custom substances
# ---------------------------------------------------------------------------

CUSTOM_SUBSTANCES_PATH = os.path.join(os.path.dirname(__file__), "custom_substances.json")
HIDDEN_SUBSTANCES_PATH = os.path.join(os.path.dirname(__file__), "hidden_substances.json")


def _load_custom_substances() -> list:
    """Load custom substances from JSON file."""
    if not os.path.exists(CUSTOM_SUBSTANCES_PATH):
        return []
    try:
        with open(CUSTOM_SUBSTANCES_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []


def _save_custom_substances(data: list):
    """Save custom substances to JSON file."""
    with open(CUSTOM_SUBSTANCES_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _load_hidden_codes() -> list:
    """Список кодов встроенных веществ, которые пользователь скрыл."""
    if not os.path.exists(HIDDEN_SUBSTANCES_PATH):
        return []
    try:
        with open(HIDDEN_SUBSTANCES_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, list):
                return [str(c) for c in data]
            return []
    except (json.JSONDecodeError, IOError):
        return []


def _save_hidden_codes(codes: list):
    with open(HIDDEN_SUBSTANCES_PATH, "w", encoding="utf-8") as f:
        json.dump(list(codes), f, ensure_ascii=False, indent=2)


def _get_all_substances() -> list:
    """Merge built-in SUBSTANCES with custom ones, исключая скрытые встроенные.
    Custom override built-in by code."""
    custom = _load_custom_substances()
    hidden = set(_load_hidden_codes())
    custom_codes = {s["code"] for s in custom}
    merged = [
        s for s in SUBSTANCES
        if s["code"] not in custom_codes and s["code"] not in hidden
    ]
    merged.extend(custom)
    return merged


class SubstanceBody(BaseModel):
    code: str
    name: str
    pdk_mr: Optional[float] = None
    pdk_ss: Optional[float] = None
    hazard_class: Optional[int] = None
    F: float = 1.0

app = FastAPI(title="ОНД-86 API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# GET /api/cities  — список городов
# ---------------------------------------------------------------------------

@app.get("/api/cities")
def get_cities():
    result = []
    for name, data in CITIES.items():
        result.append({
            "name": name,
            "lat": data["lat"],
            "lon": data["lon"],
            "A": data["A"],
            "T_warm": data.get("T_warm", 20.0),
            "T_cold": data.get("T_cold", 0.0),
            "T_avg": data.get("T_avg", 10.0),
        })
    return result


# ---------------------------------------------------------------------------
# GET /api/substances  — справочник веществ (built-in + custom merged)
# ---------------------------------------------------------------------------

@app.get("/api/substances")
def get_substances():
    return _get_all_substances()


# ---------------------------------------------------------------------------
# POST /api/substances  — добавить новое вещество
# ---------------------------------------------------------------------------

@app.post("/api/substances")
def add_substance(body: SubstanceBody):
    all_subs = _get_all_substances()
    if any(s["code"] == body.code for s in all_subs):
        raise HTTPException(status_code=409, detail=f"Вещество с кодом '{body.code}' уже существует")

    custom = _load_custom_substances()
    new_item = body.model_dump()
    new_item["custom"] = True
    custom.append(new_item)
    _save_custom_substances(custom)
    return new_item


# ---------------------------------------------------------------------------
# PUT /api/substances/{code}  — обновить вещество
# ---------------------------------------------------------------------------

@app.put("/api/substances/{code}")
def update_substance(code: str, body: SubstanceBody):
    custom = _load_custom_substances()

    # Check if it is a built-in substance being edited for the first time
    builtin = next((s for s in SUBSTANCES if s["code"] == code), None)
    existing_custom = next((s for s in custom if s["code"] == code), None)

    # If code is being changed, make sure new code doesn't conflict
    if body.code != code:
        all_subs = _get_all_substances()
        if any(s["code"] == body.code for s in all_subs if s["code"] != code):
            raise HTTPException(status_code=409, detail=f"Вещество с кодом '{body.code}' уже существует")

    updated = body.model_dump()
    updated["custom"] = True

    if existing_custom:
        # Update existing custom entry
        custom = [updated if s["code"] == code else s for s in custom]
    elif builtin:
        # Override a built-in substance — add to custom list
        custom.append(updated)
    else:
        raise HTTPException(status_code=404, detail=f"Вещество с кодом '{code}' не найдено")

    _save_custom_substances(custom)
    return updated


# ---------------------------------------------------------------------------
# DELETE /api/substances/{code}  — удалить вещество
# ---------------------------------------------------------------------------

@app.delete("/api/substances/{code}")
def delete_substance(code: str):
    custom = _load_custom_substances()
    builtin = next((s for s in SUBSTANCES if s["code"] == code), None)
    existing_custom = next((s for s in custom if s["code"] == code), None)

    if not builtin and not existing_custom:
        raise HTTPException(status_code=404, detail=f"Вещество с кодом '{code}' не найдено")

    # Удаляем кастомную запись (если была)
    if existing_custom:
        custom = [s for s in custom if s["code"] != code]
        _save_custom_substances(custom)

    # Если есть встроенный аналог с тем же кодом — добавляем в список скрытых,
    # чтобы он не возвращался обратно в результате _get_all_substances().
    if builtin:
        hidden = set(_load_hidden_codes())
        hidden.add(code)
        _save_hidden_codes(sorted(hidden))

    return {"ok": True, "deleted": code, "hidden_builtin": bool(builtin)}


# ---------------------------------------------------------------------------
# POST /api/substances/restore-defaults  — вернуть все скрытые встроенные
# ---------------------------------------------------------------------------

@app.post("/api/substances/restore-defaults")
def restore_default_substances():
    _save_hidden_codes([])
    return {"ok": True}


# ---------------------------------------------------------------------------
# POST /api/import  — импорт источников из CSV/Excel
# ---------------------------------------------------------------------------

@app.post("/api/import")
async def import_sources(file: UploadFile = File(...)):
    filename = file.filename.lower()
    content = await file.read()

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        sources = parse_csv(text)
    elif filename.endswith(".xlsx"):
        sources = parse_excel(content)
    else:
        raise HTTPException(status_code=400, detail="Поддерживаются только CSV (.csv) и Excel (.xlsx)")

    # Разделяем на валидные и с ошибками
    valid = []
    errors = []
    for s in sources:
        row_errors = s.pop("_errors", [])
        s.pop("_row", None)
        if row_errors:
            errors.extend(row_errors)
        else:
            valid.append(s)

    return {"sources": valid, "errors": errors, "total": len(sources), "valid_count": len(valid)}


# ---------------------------------------------------------------------------
# GET /api/import/template  — скачать шаблон CSV
# ---------------------------------------------------------------------------

@app.get("/api/import/template")
def get_import_template():
    xlsx_bytes = generate_template_xlsx()
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_sources.xlsx"},
    )


# ---------------------------------------------------------------------------
# GET /api/weather  — текущая погода через Open-Meteo
# ---------------------------------------------------------------------------

@app.get("/api/weather")
async def get_weather(lat: float = Query(...), lon: float = Query(...)):
    import httpx
    from datetime import datetime, timezone

    url = "https://api.open-meteo.com/v1/forecast"
    params = {
        "latitude": lat,
        "longitude": lon,
        "current": "temperature_2m,wind_speed_10m,wind_direction_10m,cloud_cover,is_day",
        "timezone": "auto",
    }

    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(url, params=params)
            resp.raise_for_status()
            data = resp.json()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Ошибка Open-Meteo: {e}")

    current = data.get("current", {})
    wind_speed = current.get("wind_speed_10m", 3.0)
    cloud_cover = current.get("cloud_cover", 50)
    is_day = current.get("is_day", 1)

    # Определение класса устойчивости по Пасквиллу–Тёрнеру
    stability_class = _determine_stability(wind_speed, cloud_cover, bool(is_day))

    return {
        "temperature": current.get("temperature_2m"),
        "wind_speed": wind_speed,
        "wind_direction": current.get("wind_direction_10m"),
        "cloud_cover": cloud_cover,
        "is_day": bool(is_day),
        "stability_class": stability_class,
        "source": "Open-Meteo",
    }


def _determine_stability(wind_speed: float, cloud_cover: float, is_day: bool) -> str:
    """Класс устойчивости по Пасквиллу–Тёрнеру."""
    if is_day:
        if cloud_cover < 30:
            radiation = "strong"
        elif cloud_cover < 70:
            radiation = "moderate"
        else:
            radiation = "weak"

        if wind_speed < 2:
            return {"strong": "A", "moderate": "A", "weak": "B"}[radiation]
        elif wind_speed < 3:
            return {"strong": "A", "moderate": "B", "weak": "C"}[radiation]
        elif wind_speed < 5:
            return {"strong": "B", "moderate": "B", "weak": "C"}[radiation]
        elif wind_speed < 6:
            return {"strong": "C", "moderate": "C", "weak": "D"}[radiation]
        else:
            return {"strong": "C", "moderate": "D", "weak": "D"}[radiation]
    else:
        if cloud_cover < 50:
            if wind_speed < 2:
                return "F"
            elif wind_speed < 3:
                return "E"
            elif wind_speed < 5:
                return "D"
            else:
                return "D"
        else:
            if wind_speed < 2:
                return "F"
            elif wind_speed < 3:
                return "F"
            elif wind_speed < 5:
                return "E"
            else:
                return "D"


def _grid_kwargs(grid) -> dict:
    """Формирует kwargs для compute_grid из GridInput (совместимость старый/новый формат)."""
    if grid.x_length is not None:
        return dict(
            x_length=grid.x_length,
            y_length=grid.y_length,
            grid_step=grid.step,
            source_offset_x=grid.source_offset_x,
            source_offset_y=grid.source_offset_y,
        )
    # Старый формат (radius)
    return dict(grid_radius=grid.radius, grid_step=grid.step)


# ---------------------------------------------------------------------------
# POST /api/calculate  — расчёт поля концентраций
# ---------------------------------------------------------------------------

@app.post("/api/calculate", response_model=CalculationResponse)
def calculate(req: CalculationRequest):
    if not req.sources:
        raise HTTPException(status_code=400, detail="Необходимо указать хотя бы один источник")

    city_data = CITIES.get(req.meteo.city)
    if city_data is None:
        # Если города нет в БД — используем коэффициент по умолчанию
        city_data = {"A": 200, "lat": req.sources[0].lat, "lon": req.sources[0].lon}

    multi = compute_per_substance(
        sources=req.sources,
        meteo=req.meteo,
        city_data=city_data,
        **_grid_kwargs(req.grid),
    )

    # Собираем плоский список SubstanceResult
    by_substance = []
    for code, data in multi["by_substance"].items():
        sub_meta = data.get("substance") or {}
        top_pts_raw = data.get("top_points") or []
        top_pts = [
            TopPoint(
                qh=round(tp["qh"], 6),
                c_mg=round(tp["c_mg"], 6),
                x_m=tp["x_m"],
                y_m=tp["y_m"],
                lat=tp["lat"],
                lon=tp["lon"],
                wind_dir_deg=tp.get("wind_dir_deg"),
                wind_speed_ms=tp["wind_speed_ms"],
                contributions=[TopPointContribution(**c) for c in tp["contributions"]],
            )
            for tp in top_pts_raw
        ]
        by_substance.append(SubstanceResult(
            code=code if code != "unknown" else (sub_meta.get("code") if sub_meta else None),
            name=sub_meta.get("name") if sub_meta else None,
            hazard_class=sub_meta.get("hazard_class") if sub_meta else None,
            pdk=data["pdk"],
            max_c=round(data["max_c"], 6),
            max_lat=data["max_lat"],
            max_lon=data["max_lon"],
            exceeds_pdk=data["exceeds_pdk"],
            ratio_to_pdk=round(data["ratio_to_pdk"], 4),
            points=[
                GridPoint(lat=p["lat"], lon=p["lon"], c=round(p["c"], 6))
                for p in data["points"]
            ],
            source_results=[SourceResult(**sr) for sr in data["source_results"]],
            top_points=top_pts,
        ))

    # Главное (худшее) вещество — поля для обратной совместимости
    primary_code = multi["primary_code"]
    primary_data = multi["by_substance"][primary_code]

    primary_points = [
        GridPoint(lat=p["lat"], lon=p["lon"], c=round(p["c"], 6))
        for p in primary_data["points"]
    ]
    primary_pdk = primary_data["pdk"]

    # Расчёт границы СЗЗ — по главному веществу, как и раньше
    center_lat = sum(s.lat for s in req.sources) / len(req.sources)
    center_lon = sum(s.lon for s in req.sources) / len(req.sources)
    import numpy as _np
    primary_lats = _np.array([p.lat for p in primary_points])
    primary_lons = _np.array([p.lon for p in primary_points])
    primary_total_mg = _np.array([p.c for p in primary_points])
    szz_data = compute_szz_boundary(primary_lats, primary_lons, primary_total_mg,
                                    primary_pdk, center_lat, center_lon)
    szz = SzzResult(
        boundary=[SzzBoundaryPoint(**p) for p in szz_data["boundary"]],
        max_distance_m=szz_data["max_distance_m"],
        min_distance_m=szz_data["min_distance_m"],
        area_ha=szz_data["area_ha"],
    ) if szz_data["max_distance_m"] > 0 else None

    return CalculationResponse(
        points=primary_points,
        max_c=round(primary_data["max_c"], 6),
        max_lat=primary_data["max_lat"],
        max_lon=primary_data["max_lon"],
        source_results=[SourceResult(**sr) for sr in primary_data["source_results"]],
        exceeds_pdk=primary_data["exceeds_pdk"],
        pdk=primary_pdk,
        szz=szz,
        by_substance=by_substance,
        primary_code=primary_code if primary_code != "unknown" else None,
    )


# ---------------------------------------------------------------------------
# POST /api/tables  — генерация таблиц ПДВ/ОВОС
# ---------------------------------------------------------------------------

@app.post("/api/tables")
def get_tables(req: CalculationRequest):
    if not req.sources:
        raise HTTPException(status_code=400, detail="Необходимо указать хотя бы один источник")

    city_data = CITIES.get(req.meteo.city, {"A": 200})

    lats, lons, total_mg, src_results, _dir_at_max = compute_grid(
        sources=req.sources,
        meteo=req.meteo,
        city_data=city_data,
        **_grid_kwargs(req.grid),
    )

    max_idx = int(total_mg.argmax())
    max_c = float(total_mg[max_idx])
    pdk = req.pdk or 0.5

    result_dict = {
        "max_c": max_c,
        "source_results": src_results,
        "exceeds_pdk": max_c > pdk,
        "pdk": pdk,
    }

    substance = req.substance.model_dump() if req.substance else None

    pdv = generate_pdv_tables(req.sources, req.meteo, city_data, substance, result_dict)
    ovos = generate_ovos_tables(req.sources, req.meteo, city_data, substance, result_dict)

    return {"pdv": pdv, "ovos": ovos}


# ---------------------------------------------------------------------------
# POST /api/export/pdf  — генерация PDF-отчёта
# ---------------------------------------------------------------------------

@app.post("/api/export/pdf")
def export_pdf(req: CalculationRequest):
    import traceback
    try:
        return _export_pdf_inner(req)
    except Exception as e:
        # Печатаем traceback в stdout — увидим в Render Logs
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка генерации PDF: {type(e).__name__}: {str(e)[:500]}"
        )


def _export_pdf_inner(req: CalculationRequest):
    request_dict = req.model_dump()

    # Если фронт прислал готовый результат /api/calculate — используем его.
    # Это пропускает дорогой compute_per_substance и снижает время экспорта
    # в 2 раза на Render free-тарифе.
    if req.precomputed_result and isinstance(req.precomputed_result, dict):
        result_dict = dict(req.precomputed_result)
        # Гарантируем минимально необходимые ключи для generate_pdf
        result_dict.setdefault("points", [])
        result_dict.setdefault("by_substance", [])
        result_dict.setdefault("pdk", req.pdk or 0.5)
    else:
        city_data = CITIES.get(req.meteo.city, {"A": 200})
        multi = compute_per_substance(
            sources=req.sources,
            meteo=req.meteo,
            city_data=city_data,
            **_grid_kwargs(req.grid),
        )
        primary_code = multi["primary_code"]
        primary = multi["by_substance"][primary_code]
        result_dict = {
            "points": primary["points"],
            "max_c": primary["max_c"],
            "max_lat": primary["max_lat"],
            "max_lon": primary["max_lon"],
            "source_results": primary["source_results"],
            "exceeds_pdk": primary["exceeds_pdk"],
            "pdk": primary["pdk"],
            "by_substance": [
                {
                    "code": code if code != "unknown" else None,
                    "substance": data.get("substance"),
                    "pdk": data["pdk"],
                    "max_c": data["max_c"],
                    "max_lat": data["max_lat"],
                    "max_lon": data["max_lon"],
                    "exceeds_pdk": data["exceeds_pdk"],
                    "ratio_to_pdk": data["ratio_to_pdk"],
                    "points": data["points"],
                    "source_results": data["source_results"],
                    "top_points": data.get("top_points") or [],
                }
                for code, data in multi["by_substance"].items()
            ],
            "primary_code": primary_code if primary_code != "unknown" else None,
        }

    pdf_bytes = generate_pdf(request_dict, result_dict)

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=ond86_report.pdf"},
    )


# ---------------------------------------------------------------------------
# POST /api/export/excel  — генерация Excel-отчёта (таблицы ПДВ + ОВОС)
# ---------------------------------------------------------------------------

@app.post("/api/export/excel")
def export_excel(req: CalculationRequest):
    city_data = CITIES.get(req.meteo.city, {"A": 200})

    lats, lons, total_mg, src_results, _dir_at_max = compute_grid(
        sources=req.sources,
        meteo=req.meteo,
        city_data=city_data,
        **_grid_kwargs(req.grid),
    )

    max_idx = int(total_mg.argmax())
    max_c = float(total_mg[max_idx])
    pdk = req.pdk or 0.5

    result_dict = {
        "max_c": max_c,
        "source_results": src_results,
        "exceeds_pdk": max_c > pdk,
        "pdk": pdk,
    }

    substance = req.substance.model_dump() if req.substance else None

    pdv = generate_pdv_tables(req.sources, req.meteo, city_data, substance, result_dict)
    ovos = generate_ovos_tables(req.sources, req.meteo, city_data, substance, result_dict)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    all_tables = {**pdv, **ovos}

    for key, table_data in all_tables.items():
        title = table_data["title"]
        columns = table_data["columns"]
        rows = table_data["rows"]

        ws = wb.create_sheet(title=key)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal="center")

        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows, 4):
            values = list(row_data.values())
            for col_idx, val in enumerate(values, 1):
                if col_idx > len(columns):
                    break
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        # Auto-width
        for col_idx in range(1, len(columns) + 1):
            max_len = len(str(columns[col_idx - 1]))
            for row in ws.iter_rows(min_row=4, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "ond86_tables.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# POST /api/export/map-png  — ZIP с прозрачными PNG карт рассеивания
# (по одному файлу на каждое вещество, для наложения в CorelDraw)
# ---------------------------------------------------------------------------

@app.post("/api/export/map-png")
def export_map_png(req: CalculationRequest):
    import zipfile
    import re
    import traceback

    try:
        # Берём готовые результаты с фронта или считаем заново
        if req.precomputed_result and isinstance(req.precomputed_result, dict):
            res = req.precomputed_result
            by_substance_data = res.get("by_substance") or []
        else:
            city_data = CITIES.get(req.meteo.city, {"A": 200})
            multi = compute_per_substance(
                sources=req.sources, meteo=req.meteo,
                city_data=city_data, **_grid_kwargs(req.grid),
            )
            by_substance_data = []
            for code, data in multi["by_substance"].items():
                sub_meta = data.get("substance") or {}
                by_substance_data.append({
                    "code": code if code != "unknown" else None,
                    "substance": sub_meta if isinstance(sub_meta, dict) else {
                        "code": getattr(sub_meta, "code", None),
                        "name": getattr(sub_meta, "name", None),
                    },
                    "pdk": data["pdk"],
                    "points": data["points"],
                })

        if not by_substance_data:
            raise HTTPException(status_code=400, detail="Нет данных для построения карт. Сначала выполните расчёт.")

        # Конвертируем источники в простые dict (Pydantic модели или dict)
        sources_for_map = []
        for src in req.sources:
            try:
                sources_for_map.append({"lat": src.lat, "lon": src.lon, "name": src.name})
            except AttributeError:
                if isinstance(src, dict):
                    sources_for_map.append({
                        "lat": src.get("lat"), "lon": src.get("lon"), "name": src.get("name", "")
                    })

        boundary = []
        if req.enterprise and req.enterprise.boundary:
            for p in req.enterprise.boundary:
                try:
                    boundary.append({"lat": p.lat, "lon": p.lon})
                except AttributeError:
                    if isinstance(p, dict):
                        boundary.append({"lat": p.get("lat"), "lon": p.get("lon")})

        grid_data = req.grid.model_dump() if req.grid else {}

        # Создаём ZIP в памяти
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for idx, sub in enumerate(by_substance_data):
                if isinstance(sub, dict):
                    # Поддерживаем две формы: name/code на верхнем уровне sub
                    # (precomputed_result от фронта) и внутри sub["substance"]
                    # (recomputed на бэке — структура из compute_per_substance).
                    sub_meta = sub.get("substance") or {}
                    nested_name = sub_meta.get("name") if isinstance(sub_meta, dict) else None
                    sub_name = (
                        nested_name
                        or sub.get("name")
                        or sub.get("code")
                        or f"вещество-{idx+1}"
                    )
                    sub_pdk = sub.get("pdk") or 0.5
                    sub_points = sub.get("points") or []
                else:
                    continue

                if not sub_points:
                    continue

                # Безопасное имя файла (общее для обоих вариантов карты)
                safe_name = re.sub(r"[^A-Za-zА-Яа-я0-9_-]+", "_", sub_name).strip("_") or f"map_{idx+1}"

                # 1) Изолинии (цветное облако) — как и раньше
                png_buf = _make_transparent_dispersion_map(
                    sub_points,
                    sources=sources_for_map,
                    boundary=boundary,
                    pdk=sub_pdk,
                    substance_name=sub_name,
                    show_axes=False,   # для CorelDraw — без осей
                    show_title=False,  # без заголовка
                    grid_data=grid_data,
                )
                if png_buf is not None:
                    zf.writestr(f"izolinii_{safe_name}.png", png_buf.getvalue())

                # 2) Сетка ОНД с числами — прозрачный фон, полупрозрачная
                #    заливка, высокое разрешение (чёткие цифры при увеличении).
                grid_step = grid_data.get("step", 500) if grid_data else 500
                try:
                    grid_buf = _make_concentration_plot(
                        sub_points,
                        grid_step=grid_step,
                        title=f"Карта рассеивания {sub_name}, доли ПДК",
                        grid_params=grid_data,
                        sources=None,
                        boundary=boundary,
                        pdk=sub_pdk,
                        transparent=True,
                        dpi=300,
                    )
                    if grid_buf is not None:
                        zf.writestr(f"setka-ond_{safe_name}.png", grid_buf.getvalue())
                except Exception as grid_err:
                    # Сетка не критична — изолинии уже в архиве, не валим весь экспорт
                    print(f"[map-png] Сетка ОНД для «{sub_name}» не построена: {grid_err}")

        zip_buf.seek(0)
        return StreamingResponse(
            zip_buf,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=karty-rasseivaniya.zip"},
        )

    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка генерации карт PNG: {type(e).__name__}: {str(e)[:500]}"
        )


# ---------------------------------------------------------------------------
# GET /api/health  — health check (для облачных пробок Render/Railway)
# ---------------------------------------------------------------------------

@app.get("/api/health")
def health():
    return {"status": "ok", "app": "ОНД-86 Расчёт рассеивания"}


# ---------------------------------------------------------------------------
# Раздача собранного фронтенда (frontend/dist) для облачного развертывания.
# В dev-режиме папки dist может не быть — тогда отдаём JSON-заглушку,
# а Vite крутится отдельно на :5173 с прокси /api → :8000.
# ---------------------------------------------------------------------------

from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

_BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DIST_DIR = os.path.abspath(os.path.join(_BACKEND_DIR, "..", "frontend", "dist"))
_INDEX_HTML = os.path.join(_DIST_DIR, "index.html")


if os.path.isdir(_DIST_DIR):
    # Раздаём ассеты Vite (бандлы JS/CSS), они с хэшем в имени —
    # достаточно прямой StaticFiles-mount.
    _ASSETS_DIR = os.path.join(_DIST_DIR, "assets")
    if os.path.isdir(_ASSETS_DIR):
        app.mount("/assets", StaticFiles(directory=_ASSETS_DIR), name="assets")

    # Catch-all: всё, что не /api/* и не /assets/* — пытаемся отдать как файл из dist,
    # иначе возвращаем index.html (для SPA-роутинга).
    @app.get("/{full_path:path}", include_in_schema=False)
    def serve_spa(full_path: str):
        if full_path.startswith("api/") or full_path.startswith("assets/"):
            raise HTTPException(status_code=404)
        candidate = os.path.join(_DIST_DIR, full_path) if full_path else _INDEX_HTML
        if full_path and os.path.isfile(candidate):
            return FileResponse(candidate)
        if os.path.isfile(_INDEX_HTML):
            return FileResponse(_INDEX_HTML)
        raise HTTPException(status_code=404)
else:
    # Frontend ещё не собран — отдаём JSON по корню для health-check.
    @app.get("/")
    def root():
        return JSONResponse({
            "status": "ok",
            "app": "ОНД-86 Расчёт рассеивания",
            "note": "frontend/dist not built — use Vite dev server on :5173",
        })
