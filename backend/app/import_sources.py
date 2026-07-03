"""
Импорт источников из CSV и Excel файлов.
"""

import csv
import io
from typing import List, Dict, Any

# Маппинг допустимых названий столбцов к каноническим
COLUMN_MAP = {
    # Название / номер
    "название": "name", "name": "name", "источник": "name", "source": "name",
    "№": "name", "номер": "name", "number": "name", "n": "name",
    # Код вещества
    "код вещества": "substance_code", "код": "substance_code",
    "substance_code": "substance_code", "code": "substance_code",
    # Название вещества
    "название вещества": "substance_name", "вещество": "substance_name",
    "substance": "substance_name", "substance_name": "substance_name",
    # Высота
    "высота": "height", "h": "height", "h, м": "height", "height": "height",
    "высота трубы": "height", "высота, м": "height",
    "высота (h), м": "height", "высота (h)": "height",
    # Диаметр
    "диаметр": "diameter", "d": "diameter", "d, м": "diameter", "diameter": "diameter",
    "диаметр устья": "diameter",
    "диаметр (d), м": "diameter", "диаметр (d)": "diameter",
    # Скорость
    "скорость": "velocity", "w0": "velocity", "w0, м/с": "velocity", "velocity": "velocity",
    "скорость выхода": "velocity",
    "скорость (w0), м/с": "velocity", "скорость (w0)": "velocity",
    # Температура
    "температура": "temperature", "tг": "temperature", "tг, °c": "temperature",
    "temperature": "temperature", "температура газов": "temperature",
    "температура (tг), °c": "temperature", "температура (tг)": "temperature",
    # Выброс г/с
    "выброс г/с": "emission_gs", "m": "emission_gs", "m, г/с": "emission_gs",
    "emission_gs": "emission_gs", "г/с": "emission_gs", "выброс": "emission_gs",
    "выброс (m), г/с": "emission_gs", "выброс (m)": "emission_gs",
    # Выброс т/год
    "выброс т/год": "emission_ty", "m год": "emission_ty", "m, т/год": "emission_ty",
    "emission_ty": "emission_ty", "т/год": "emission_ty",
    "выброс, т/год": "emission_ty",
    # Координаты
    "широта": "lat", "lat": "lat", "latitude": "lat",
    "долгота": "lon", "lon": "lon", "longitude": "lon", "lng": "lon",
}

REQUIRED_FIELDS = {"name", "height", "diameter", "velocity", "temperature", "emission_gs"}


def _normalize_header(header: str) -> str:
    """Приводит заголовок к каноническому имени."""
    h = header.strip().lower().replace("°", "°")
    return COLUMN_MAP.get(h, h)


def parse_csv(content: str) -> List[Dict[str, Any]]:
    """Парсит CSV-текст и возвращает список источников.

    Поддерживает «протяжку» параметров трубы из строки выше:
    если у текущей строки заполнены только Код/выбросы, а параметры трубы
    пусты — они наследуются от предыдущей валидной строки того же источника.
    """
    # Определяем разделитель
    delimiter = ";"
    if content.count(",") > content.count(";"):
        delimiter = ","

    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    rows = list(reader)

    if len(rows) < 2:
        return []

    headers = [_normalize_header(h) for h in rows[0]]
    sources = []

    NUMERIC_FIELDS = (
        "height", "diameter", "velocity", "temperature",
        "emission_gs", "emission_ty", "lat", "lon",
    )
    INHERITABLE_FIELDS = ("name", "height", "diameter", "velocity",
                          "temperature", "lat", "lon")

    last_inheritable: Dict[str, Any] = {}

    for row_idx, row in enumerate(rows[1:], start=2):
        if len(row) < len(headers):
            row += [""] * (len(headers) - len(row))

        source = {}
        errors = []
        any_numeric_parsed = False

        for i, header in enumerate(headers):
            val = row[i].strip() if i < len(row) else ""

            if header == "name":
                source[header] = val or None
            elif header in ("substance_code", "substance_name"):
                source[header] = val or None
            elif header in NUMERIC_FIELDS:
                if not val:
                    source[header] = None
                    continue
                try:
                    source[header] = float(val.replace(",", "."))
                    any_numeric_parsed = True
                except ValueError:
                    source[header] = None
                    errors.append(f"Строка {row_idx}, '{headers[i]}': нечисловое значение '{val}'")

        # Наследуем параметры трубы из строки выше
        for f in INHERITABLE_FIELDS:
            if (source.get(f) is None) and last_inheritable.get(f) is not None:
                source[f] = last_inheritable[f]

        has_emission = (
            source.get("emission_gs") is not None
            or source.get("emission_ty") is not None
        )

        # Если в строке вообще нет числовых данных и нет выбросов —
        # это пустая/служебная строка (хвост, «Итого», комментарий).
        if not any_numeric_parsed and not has_emission:
            continue

        # Обновляем «последние известные» параметры трубы
        for f in INHERITABLE_FIELDS:
            if source.get(f) is not None:
                last_inheritable[f] = source[f]

        if not source.get("name"):
            source["name"] = last_inheritable.get("name") or f"Источник {row_idx - 1}"
            last_inheritable["name"] = source["name"]

        # Проверка обязательных полей
        for field in REQUIRED_FIELDS:
            if field not in source or source[field] is None:
                if field == "name":
                    source["name"] = f"Источник {row_idx - 1}"
                else:
                    errors.append(f"Строка {row_idx}: отсутствует обязательное поле '{field}'")

        source["_row"] = row_idx
        source["_errors"] = errors
        sources.append(source)

    return sources


def parse_excel(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Парсит Excel (.xlsx) файл и возвращает список источников.

    Поддерживает «карточный» вид с merge_cells по параметрам трубы:
    в строках, относящихся к дополнительным веществам того же источника,
    значения параметров трубы — None (Excel хранит merge так, что значение
    видно только в top-left ячейке). Такие пропуски наследуются от
    предыдущей валидной строки того же источника.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [_normalize_header(str(h or "")) for h in rows[0]]
    sources = []

    NUMERIC_FIELDS = (
        "height", "diameter", "velocity", "temperature",
        "emission_gs", "emission_ty", "lat", "lon",
    )
    # Параметры трубы — это поля, которые могут быть объединены merge_cells
    # по нескольким веществам одного источника. Их «протягиваем» из строки выше.
    INHERITABLE_FIELDS = ("name", "height", "diameter", "velocity",
                          "temperature", "lat", "lon")

    last_inheritable: Dict[str, Any] = {}

    for row_idx, row in enumerate(rows[1:], start=2):
        source = {}
        errors = []
        any_numeric_parsed = False

        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None

            if header == "name":
                source[header] = str(val).strip() if val else None
            elif header in ("substance_code", "substance_name"):
                # Excel часто хранит код вещества числом: "0301" → 301 или 301.0.
                # Возвращаем строку без ".0"; ведущие нули восстановить нельзя —
                # их учитывает сопоставление на фронтенде.
                if val is None or val == "":
                    source[header] = None
                elif isinstance(val, float) and val.is_integer():
                    source[header] = str(int(val))
                else:
                    source[header] = str(val).strip()
            elif header in NUMERIC_FIELDS:
                if val is None or val == "":
                    source[header] = None
                    continue
                try:
                    # Excel в русской локали иногда хранит числа как строки
                    # с запятой "0,000000205". Заменяем запятую на точку и
                    # удаляем неразрывные пробелы (тысячный разделитель).
                    if isinstance(val, str):
                        cleaned = val.strip().replace("\xa0", "").replace(" ", "")
                        cleaned = cleaned.replace(",", ".")
                        source[header] = float(cleaned) if cleaned else None
                        if source[header] is not None:
                            any_numeric_parsed = True
                    else:
                        source[header] = float(val)
                        any_numeric_parsed = True
                except (ValueError, TypeError):
                    source[header] = None
                    errors.append(f"Строка {row_idx}, '{headers[i]}': нечисловое значение «{val}»")

        # Наследуем параметры трубы из строки выше, если пусто
        for f in INHERITABLE_FIELDS:
            if (source.get(f) is None) and last_inheritable.get(f) is not None:
                source[f] = last_inheritable[f]

        # Признак «строка содержит данные о выбросе» — хотя бы один из г/с / т/год
        has_emission = (
            source.get("emission_gs") is not None
            or source.get("emission_ty") is not None
        )

        # Пустая или служебная строка — тихо пропускаем.
        # Считаем строку служебной, если:
        #   - вообще нет числовых данных,
        #   - И нет выбросов, унаследованных параметров недостаточно для источника.
        if not any_numeric_parsed and not has_emission:
            continue

        # Обновляем «последние известные» параметры трубы для следующих строк
        for f in INHERITABLE_FIELDS:
            if source.get(f) is not None:
                last_inheritable[f] = source[f]

        # Имя по умолчанию
        if not source.get("name"):
            source["name"] = last_inheritable.get("name") or f"Источник {row_idx - 1}"
            last_inheritable["name"] = source["name"]

        for field in REQUIRED_FIELDS:
            if field not in source or source[field] is None:
                if field == "name":
                    source["name"] = f"Источник {row_idx - 1}"
                else:
                    errors.append(f"Строка {row_idx}: отсутствует обязательное поле '{field}'")

        source["_row"] = row_idx
        source["_errors"] = errors
        sources.append(source)

    wb.close()
    return sources


def generate_template_csv() -> str:
    """Генерирует CSV-шаблон для импорта (с BOM для Excel).

    Каждая строка — одно вещество от одного источника.
    Если у одной трубы несколько веществ — повторите строку с тем же
    "Названием" и параметрами трубы, поменяв только колонки вещества и выбросов.
    """
    header = "Название;Код вещества;Высота (H), м;Диаметр (D), м;Скорость (w0), м/с;Температура (Tг), °C;Выброс (M), г/с;Выброс, т/год;Широта;Долгота"
    examples = "\n".join([
        "Труба ТЭЦ-1;0301;45;1.2;12.0;180;8.5;268.06;41.2995;69.2401",
        "Труба ТЭЦ-1;0337;45;1.2;12.0;180;12.0;378.43;41.2995;69.2401",
        "Труба ТЭЦ-1;0330;45;1.2;12.0;180;3.5;110.4;41.2995;69.2401",
    ])
    return f"{header}\n{examples}\n"


def generate_template_xlsx() -> bytes:
    """Генерирует Excel-шаблон для импорта.

    Плоская таблица: одна строка = одно вещество от одного источника.
    Параметры трубы (Название, H, D, w₀, Tг, Lat, Lon) визуально
    объединены (merge) по нескольким строкам одного источника —
    в каждой подстроке кода/выбросов параметры трубы повторяются
    зрительно как «карточка».
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Источники"

    headers = [
        "Название", "Код вещества", "Высота (H), м", "Диаметр (D), м",
        "Скорость (w0), м/с", "Температура (Tг), °C",
        "Выброс (M), г/с", "Выброс, т/год", "Широта", "Долгота",
    ]
    # Индексы колонок, которые относятся к «параметрам трубы» (1-based).
    # Их объединяем merge_cells по всем строкам одного источника.
    SRC_PARAM_COLS = (1, 3, 4, 5, 6, 9, 10)
    # Колонки «по веществу» (свои значения в каждой строке)
    SUBSTANCE_COLS = (2, 7, 8)

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    src_param_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    thin = Side(style="thin", color="CBD5E1")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = thin_border
    ws.row_dimensions[1].height = 36

    # Демо-данные: 3 источника с разным числом веществ
    sources_demo = [
        {
            "name": "Труба ТЭЦ-1",
            "height": 45, "diameter": 1.2, "velocity": 12.0,
            "temperature": 180, "lat": 41.2995, "lon": 69.2401,
            "emissions": [
                ("0301", 8.5, 268.06),
                ("0337", 12.0, 378.43),
                ("0330", 3.5, 110.40),
            ],
        },
        {
            "name": "Труба котельной №2",
            "height": 25, "diameter": 0.8, "velocity": 8.5,
            "temperature": 140, "lat": 41.3010, "lon": 69.2480,
            "emissions": [
                ("0301", 2.1, 66.20),
                ("2902", 0.8, 25.20),
            ],
        },
        {
            "name": "Вентшахта",
            "height": 8, "diameter": 0.56, "velocity": 2.4,
            "temperature": 28.5, "lat": 41.3045, "lon": 69.2510,
            "emissions": [
                ("0123", 0.053, 0.06869),
                ("0143", 0.000018, 0.0000024),
            ],
        },
    ]

    row = 2
    for src in sources_demo:
        n = len(src["emissions"])
        first_row = row
        last_row = row + n - 1

        # Записываем параметры трубы в первую строку «карточки»
        src_param_values = {
            1: src["name"],
            3: src["height"],
            4: src["diameter"],
            5: src["velocity"],
            6: src["temperature"],
            9: src["lat"],
            10: src["lon"],
        }
        for col, val in src_param_values.items():
            cell = ws.cell(row=first_row, column=col, value=val)
            cell.font = Font(size=10, color="1E293B")
            cell.alignment = centered
            cell.fill = src_param_fill
            cell.border = thin_border
        # Объединяем колонки параметров трубы по всем строкам источника
        if n > 1:
            for col in SRC_PARAM_COLS:
                ws.merge_cells(start_row=first_row, start_column=col,
                                end_row=last_row, end_column=col)
                # Прокрашиваем объединённую область (для merge цвет нужен у top-left)
                for r in range(first_row, last_row + 1):
                    ws.cell(row=r, column=col).fill = src_param_fill
                    ws.cell(row=r, column=col).border = thin_border

        # Строки веществ
        for i, (code, gs, ty) in enumerate(src["emissions"]):
            r = first_row + i
            for col, val in ((2, code), (7, gs), (8, ty)):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = Font(size=10, color="1E293B")
                cell.alignment = centered
                cell.border = thin_border
            ws.row_dimensions[r].height = 20

        row = last_row + 1

    # Подсказка снизу
    note = ws.cell(row=row + 1, column=1, value=(
        "Каждое вещество — отдельная строка. У одного источника параметры "
        "трубы (Название, H, D, w0, Tг, Широта, Долгота) визуально объединены: "
        "при добавлении строк просто повторяйте «Название» или оставляйте "
        "объединение — парсер унаследует параметры из строки выше."
    ))
    note.font = Font(italic=True, size=10, color="64748B")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=len(headers))
    ws.row_dimensions[row + 1].height = 36

    # Ширины колонок
    col_widths = [22, 14, 14, 14, 16, 18, 16, 16, 12, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
